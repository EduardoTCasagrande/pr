import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
from datetime import date

def adicionar_venda():
    vendas = []
    data_atual = date.today()
    dia = data_atual.day  
    dia -= 1

    data_em_texto = data_atual.strftime(f'{dia:02d}/{data_atual.month:02d}')
    print(data_em_texto)

    hora = input("Digite a hora da venda (HH:MM): ")
    nome_vendedora = input("Digite o nome da vendedora: ")
    
    while True:
        tipo_produto = input("Digite o tipo do produto (B, BO, A): ").upper()
        if tipo_produto == 'A':
            tipo_produto = 'ACESSÓRIO'
        elif tipo_produto == 'B':
            tipo_produto = 'BONECA'
        else:
            tipo_produto = 'BONECO'
        nome_produto = input("Digite o nome do produto: ")
        
        venda = {
            "Tipo": tipo_produto,
            "Produto": nome_produto,
            "Hora": hora,
            "Dia": data_em_texto,
            "Vendedora": nome_vendedora
        }
        vendas.append(venda)
        
        continuar = input("Adicionar mais itens a esta venda? (s/n): ")
        if continuar.lower() != 's':
            break

    return vendas

def salvar_vendas(vendas, nome_arquivo="vendas.xlsx"):
    df = pd.DataFrame(vendas)
    
    with pd.ExcelWriter(nome_arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Vendas')
        workbook = writer.book
        worksheet = writer.sheets['Vendas']
        
        # Define o estilo da borda inferior
        thin_border = Border(bottom=Side(style='thin'))

        # Identifica as linhas a serem mescladas
        start_row = 2  # Primeira linha de dados no Excel
        for _, grupo in df.groupby(['Dia', 'Hora', 'Vendedora']):
            end_row = start_row + len(grupo) - 1
            if len(grupo) > 1:
                worksheet.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)
                worksheet.merge_cells(start_row=start_row, start_column=5, end_row=end_row, end_column=5)
                worksheet.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
                
                for col in [4, 5, 6]:
                    worksheet.cell(row=start_row, column=col).alignment = Alignment(vertical='center', horizontal='center')

            # Adiciona a linha horizontal no final do grupo
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=end_row, column=col).border = thin_border
            
            start_row = end_row + 1

    print(f"Vendas salvas no arquivo {nome_arquivo}")

vendas = []

while True:
    venda = adicionar_venda()
    vendas.extend(venda)
    
    continuar = input("Deseja adicionar outra venda? (s/n): ")
    if continuar.lower() != 's':
        break

salvar_vendas(vendas)
